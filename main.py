from vna import VNAController
import struct, traceback, socket, time
from openpyxl import Workbook#, load_workbook
from openpyxl.chart import LineChart, Reference
import numpy as np
from scipy.signal import savgol_filter
from scipy.interpolate import interp1d
import pandas as pd

def set_antenna_uniform_volt(voltage, aim_ip):
    message = bytearray([0] * 7)
    message.append(0x79)  # 118
    message.extend(f"0000".encode('utf-8'))
    voltage_float = struct.pack(">I", int(float(voltage) * 1000))
    message.extend(voltage_float)
    # RX
    first_message = message
    # Send the first message
    send_to_aim(first_message, aim_ip)

    # Tx
    second_message = message
    second_message[7] = 0x7A

    # Send the second message
    send_to_aim(second_message, aim_ip)

def send_to_aim(message, aim_ip):
    client_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    retries = 6
    while retries > 0:
        try:
            client_socket.connect((aim_ip, 1000))  # Replace with the IP of your microcontroller and the port
            break
        except:
            print(f"Connection to aim failed, retrying {retries} more times")
            retries -= 1

    client_socket.send(message)

    data = client_socket.recv(1024)

    client_socket.close()
    return data

ip = '192.168.1.21'
main_delay = 15
zero_delay = 60
start_fr = 10.7
end_fr = 12.7
volt_max = 10
volt_int = 0.1
set_antenna_uniform_volt(0, ip)
def choose(text, *options):
    # If a single list is passed, unpack it
    if len(options) == 1 and isinstance(options[0], list):
        options = options[0]
    else:
        options = list(options)

    while (user_input := input(f'{text}\n({"/".join(options)})\n')) not in options:
        print('Invalid input\n')
    return user_input

vna_controller = VNAController(13)
vna_controller.set_channel('s12')
ip = '192.168.1.21'
# -----------------SETUP START-----------------

user_input = choose('Change start and end of measured frequency range?\n'
                    f'Start frequency: {start_fr}GHz, end frequency: {end_fr}GHz', 'y','n')
if user_input == 'y':
    while True:
        try:
            start_fr = float(input('Input start frequency\nGHz'))
            if start_fr <= 0:
                raise ValueError('Start frequency must be more than zero\n')
        except:
            print('Invalid input\n')
            continue
        break

    while True:
        try:
            end_fr = float(input('Input end frequency\nGHz'))
            if end_fr <= 0:
                raise ValueError('End frequency must be more than zero\n')
        except:
            print('Invalid input\n')
            continue
        if end_fr <= start_fr:
            print('End frequency must be greater than start frequency\n')
            continue
        break

frequencies = []
while True:
    print(f'Chosen frequencies:{frequencies}')
    if len(frequencies) > 4:
        user_input = choose('Confirm? Input nothing to continue, input "-" to remove last.', '','-')
        if user_input == '-' and len(frequencies) > 0:
            frequencies.pop()
            continue
        else:
            break
    user_input = input('Input up to 5 desired frequencies, input empty to finish, input "-" to remove previous\nGHz')
    if user_input == '' and frequencies == []:
        print('Frequency list is empty\n')
        continue
    if user_input == '':
        break
    if user_input == '-' and len(frequencies) > 0:
        frequencies.pop()
        continue
    try:
        user_input = float(user_input)
        if user_input <= 0:
            raise ValueError('Frequency must be more than zero\n')
    except:
        print('Invalid input\n')
        continue
    if user_input > end_fr or user_input < start_fr:
        print(f'Frequency must be within {start_fr}GHz - {end_fr}GHz\n')
        continue

    frequencies.append(user_input)

types = choose('What is measured?\n1 - phase, 2 - gain, 3 - both', '1', '2', '3')
if types == '1':
    types = ['phase']
elif types == '2':
    types = ['gain']
else:
    types = ['gain', 'phase']

user_input = choose(f'Change aim ip ({ip})?', 'y','n')
if user_input == 'y':
    ip = input('Input aim ip\n')
set_antenna_uniform_volt(0, ip)

user_input = choose('Change delays between voltage changes?\n'
                    f'Step delay: {main_delay}s, zero volt delay: {zero_delay}s', 'y','n')
if user_input == 'y':
    while True:
        try:
            main_delay = float(input('Input how long to wait between voltage changes each step\ns'))
            if main_delay <= 0:
                raise ValueError('Time must be more than zero\n')
        except:
            print('Invalid input\n')
            continue
        break

    while True:
        try:
            zero_delay = float(input('Input how long to wait before a new 0 volt measurement\ns'))
            if zero_delay <= 0:
                raise ValueError('Time must be more than zero\n')
        except:
            print('Invalid input\n')
            continue
        break

user_input = choose('Change maximum voltage and how much voltage changes each step?\n'
                    f'Maximum voltage: {volt_max}V, voltage step: {volt_int}V', 'y','n')
if user_input == 'y':
    while True:
        try:
            volt_max = float(input('Input maximum voltage\nV'))
            if volt_max <= 0:
                raise ValueError('Voltage interval must be more than zero\n')
        except:
            print('Invalid input\n')
            continue
        break

    while True:
        try:
            volt_int = float(input('Input how much voltage changes each step\nV'))
            if volt_int <= 0:
                raise ValueError('Voltage step must be more than zero\n')
        except:
            print('Invalid input\n')
            continue
        break

file_name = input('Input excel file name\n')
#-----------------SETUP END-----------------

try:
    wb = Workbook()
    # wb = load_workbook(f'{file_name}.xlsx')
    ws = wb['Sheet']
    ws.cell(row=1, column=1, value="Frequencies")
    i = 2
    for freq in frequencies:
        ws.cell(row=i, column=1, value=freq)
        i += 1

    ws.cell(row=1, column=2, value="Measured values")
    i = 2
    for meas_type in types:
        ws.cell(row=i, column=2, value=meas_type)
        i += 1

    ws.cell(row=1, column=3, value="Max voltage")
    ws.cell(row=2, column=3, value=volt_max)
    ws.cell(row=1, column=4, value="Voltage interval")
    ws.cell(row=2, column=4, value=volt_int)

    wb.save(f'{file_name}.xlsx')
    meas_state = False

    col_i = 6
    for meas_type in types:
        print(f'Started measurement of {meas_type}')
        if meas_type == 'phase':
            meas_type = 'PHAS'
        else:
            meas_type = 'LOGM'
        vna_controller.set_ampl_or_phase(meas_type)
        vna_controller.set_frequency(start_fr, end_fr, 'GHz')
        ws.cell(row=1, column=5, value="voltage, v")
        for n, freq in enumerate(frequencies):
            ws.cell(row=1, column=col_i+n, value=f'{meas_type}, {freq}GHz')
            n += 1
            vna_controller.put_marker(n, freq)

        prev_value_dict = {}
        value_dict = {}
        offset = 0
        i = 2
        state = False
        for voltage in np.arange(0, volt_max + volt_int, volt_int):
            voltage = round(voltage, 4)
            print(f'At voltage {voltage}')
            if ws.cell(row=i, column=5).value is None:
                ws.cell(row=i, column=5, value=voltage)
            elif ws.cell(row=i, column=5).value != voltage:
                print(f"Mismatch of voltage, was {ws.cell(row=i, column=5).value}, now {voltage}")
            elif ws.cell(row=i, column=5).value == voltage:
                pass
            else:
                print(f'Weird voltage, was {ws.cell(row=i, column=5).value}, now {voltage}, rewriting.')
                ws.cell(row=i, column=5, value=voltage)
            if voltage == 0:
                set_antenna_uniform_volt(voltage, ip)
                time.sleep(zero_delay)
                if meas_type == 'PHAS':
                    vna_controller.set_data_to_memory()
                    time.sleep(5)
            else:
                set_antenna_uniform_volt(voltage, ip)
                time.sleep(main_delay)

            for n, freq in enumerate(frequencies):
                prev_value_dict[n] = value_dict.get(n, None)
                n_mark = n + 1
                value_dict[n] = vna_controller.get_mark_value(n_mark)
                if prev_value_dict.get(n, None) is None:
                    state = True
                    prev_value_dict[n] = value_dict[n]
                if meas_type == 'PHAS':
                    value_dict[n] += offset
                    if state:
                        prev_value_dict[n] += offset
                        state = False
                    if value_dict[n] - prev_value_dict[n] > 200:
                        offset += -360
                        value_dict[n] += -360
                    elif value_dict[n] - prev_value_dict[n] < -200:
                        offset += 360
                        value_dict[n] += 360
                print(f'Frequency: {freq}, {meas_type}: {value_dict[n]}')

                ws.cell(row=i, column=col_i+n, value=value_dict[n])
            i += 1
            wb.save(f'{file_name}.xlsx')
        # Create a chart
        chart = LineChart()
        chart.title = f"{meas_type} vs Voltage"
        chart.x_axis.title = "Voltage"
        chart.y_axis.title = meas_type

        # Define data range (excluding header)
        categories = Reference(ws, min_col=5, min_row=2, max_col=5, max_row=i - 1)  # voltage values
        for n, freq in enumerate(frequencies):
            col_n = col_i + n
            data = Reference(ws, min_col=col_n, min_row=1, max_col=col_n, max_row=i - 1)  # meas_type values
            chart.add_data(data, titles_from_data=True)

            if meas_type == 'PHAS':
                df = pd.read_excel(f'{file_name}.xlsx')
                phase_smooth = df[f'{meas_type}, {freq}GHz']
                volt_list = list(df['voltage, v'])
                # Smooth the phase values (replace window_length/polyorder if needed)
                phase_smooth.iloc[0] = 0
                phase_smooth = savgol_filter(phase_smooth, window_length=5, polyorder=2)
                phase_smooth = -abs(phase_smooth)
                phase_smooth[0] = 0

                # Determine target phase range
                phase_span = np.max(phase_smooth) - np.min(phase_smooth)
                if phase_span < 360:
                    # Pad the lowest phase values with 0 until span reaches 360
                    min_phase = np.min(phase_smooth)
                    padding_needed = 361 - phase_span
                    phase_smooth = np.concatenate(
                        [phase_smooth, np.full(int(padding_needed), min_phase - np.arange(1, int(padding_needed) + 1))])
                    volt_list = np.concatenate([volt_list, np.full(int(padding_needed), np.max(volt_list))])

                max_voltage_idx = np.argmax(volt_list)
                if phase_span < 360:
                    end_phase = 0
                else:
                    end_phase = int(phase_smooth[max_voltage_idx] + 360)

                # New phase points (0 to -360 step -1)
                xq_tx_44 = np.arange(end_phase, end_phase - 361, -1)
                # Interpolation: treating smoothed phase as "x", voltage as "y"
                interp_func = interp1d(phase_smooth, volt_list, kind='linear', bounds_error=False, fill_value=np.nan)
                vq_tx_44 = interp_func(xq_tx_44)
                # Stack phase and voltage
                xq_tx_44 -= np.max(xq_tx_44)
                res_tx_44 = np.column_stack((-xq_tx_44, vq_tx_44))
                # Save to CSV
                pd.DataFrame(res_tx_44, columns=['Phase', 'Voltage']) \
                    .round({'Voltage': 4}) \
                    .to_csv(f'{freq}_cal_curve.csv', index=False, header=False)



        col_i += n + 1
        chart.varyColors = False
        chart.set_categories(categories)

        # Position chart below the data (or adjust as needed)
        if meas_state == True:
            ws.add_chart(chart, f"Z2")
        else:
            ws.add_chart(chart, f"M2")
        meas_state = True

        # Save workbook with chart
        wb.save(f'{file_name}.xlsx')
except Exception as e:
    print(f"Error: {e}\nType: {type(e).__name__}\n")
    traceback.print_exc()

# set_antenna_uniform_volt(0, ip)
input("Press Enter to exit\n")
